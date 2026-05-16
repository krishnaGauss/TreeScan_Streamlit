import csv
import io
import zipfile
from functools import lru_cache

import openpyxl
import segno
from PIL import Image, ImageDraw

from constants import (
    TEMPLATE_PATH, TEXT_COLOR, QR_CX,
    QR_BOX_X1, QR_BOX_Y1, QR_BOX_X2, QR_BOX_Y2,
    NUM_Y1, NUM_Y2,
    NAME_Y1, NAME_Y2,
    ENG_NAME_Y1, ENG_NAME_Y2,
    HINDI_NAME_Y1, HINDI_NAME_Y2,
    SERIF_FONT, DEVA_FONT,
)
from drawing import fit_and_draw, parse_names


@lru_cache(maxsize=1)
def _template() -> Image.Image:
    return Image.open(TEMPLATE_PATH).convert("RGB")


def make_qr_card(
    url: str,
    english_name: str = "",
    hindi_name: str = "",
    tree_number: str = "",
) -> Image.Image:
    qr = segno.make(url, error="h")
    buf = io.BytesIO()
    qr.save(buf, kind="png", scale=10, border=2, dark="#000000", light="#ffffff")
    buf.seek(0)
    qr_img = Image.open(buf).convert("RGB")

    box_w   = QR_BOX_X2 - QR_BOX_X1
    box_h   = QR_BOX_Y2 - QR_BOX_Y1
    pad     = 20
    qr_size = min(box_w - 2 * pad, box_h - 2 * pad)

    canvas = _template().copy()
    canvas.paste(
        qr_img.resize((qr_size, qr_size), Image.LANCZOS),
        (QR_BOX_X1 + (box_w - qr_size) // 2, QR_BOX_Y1 + (box_h - qr_size) // 2),
    )

    draw = ImageDraw.Draw(canvas)

    if tree_number.strip():
        zone_h = NUM_Y2 - NUM_Y1
        fit_and_draw(
            draw, tree_number.strip(),
            cx=QR_CX, cy=(NUM_Y1 + NUM_Y2) // 2,
            max_w=int(canvas.width * 0.55), max_h=int(zone_h * 0.70),
            start_size=int(zone_h * 0.55),
            color=TEXT_COLOR,
        )

    eng = english_name.strip()
    hin = hindi_name.strip()

    if eng and hin:
        eng_h = ENG_NAME_Y2 - ENG_NAME_Y1
        fit_and_draw(
            draw, eng,
            cx=QR_CX, cy=(ENG_NAME_Y1 + ENG_NAME_Y2) // 2,
            max_w=int(canvas.width * 0.62), max_h=int(eng_h * 0.80),
            start_size=int(eng_h * 0.75),
            color=TEXT_COLOR, font_path=SERIF_FONT,
        )
        hin_h = HINDI_NAME_Y2 - HINDI_NAME_Y1
        fit_and_draw(
            draw, hin,
            cx=QR_CX, cy=(HINDI_NAME_Y1 + HINDI_NAME_Y2) // 2,
            max_w=int(canvas.width * 0.62), max_h=int(hin_h * 0.75),
            start_size=int(hin_h * 0.70),
            color=TEXT_COLOR, font_path=DEVA_FONT,
        )
    elif eng:
        zone_h = NAME_Y2 - NAME_Y1
        fit_and_draw(
            draw, eng,
            cx=QR_CX, cy=(NAME_Y1 + NAME_Y2) // 2,
            max_w=int(canvas.width * 0.62), max_h=int(zone_h * 0.70),
            start_size=int(zone_h * 0.65),
            color=TEXT_COLOR, font_path=SERIF_FONT,
        )
    elif hin:
        zone_h = NAME_Y2 - NAME_Y1
        fit_and_draw(
            draw, hin,
            cx=QR_CX, cy=(NAME_Y1 + NAME_Y2) // 2,
            max_w=int(canvas.width * 0.62), max_h=int(zone_h * 0.70),
            start_size=int(zone_h * 0.65),
            color=TEXT_COLOR, font_path=DEVA_FONT,
        )

    return canvas


def _img_to_bytes(img: Image.Image, fmt: str) -> bytes:
    buf = io.BytesIO()
    img.save(buf, format=fmt)
    buf.seek(0)
    return buf.read()


def _read_rows(file_bytes: bytes, filename: str) -> list[dict]:
    if filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = [
            {headers[i]: (str(cell.value) if cell.value is not None else "")
             for i, cell in enumerate(row)}
            for row in ws.iter_rows(min_row=2)
        ]
        wb.close()
        return rows
    return list(csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig"))))


def process_csv(
    csv_bytes: bytes, base_url: str, fmt: str, filename: str = "data.csv", on_progress=None
) -> tuple[bytes, list[str], list[str], list[str]]:
    ext      = "jpg" if fmt == "JPEG" else "png"
    base_url = base_url.rstrip("/")
    rows     = _read_rows(csv_bytes, filename)
    total    = len(rows)
    out_buf  = io.BytesIO()
    failed_no_sno:        list[str] = []
    failed_empty_vern:    list[str] = []
    failed_other:         list[str] = []

    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zout:
        for i, row in enumerate(rows):
            s_no = row.get("S_No_", "").strip()
            if not s_no:
                failed_no_sno.append(f"row {i + 2}")
            else:
                eng, hin = parse_names(row.get("Vernacular", ""))
                if not eng and not hin:
                    failed_empty_vern.append(s_no)
                    if on_progress:
                        on_progress(i + 1, total)
                    continue
                try:
                    url = f"{base_url}/{s_no}.pdf"
                    img = make_qr_card(url, eng, hin, s_no)
                    zout.writestr(f"{s_no}.{ext}", _img_to_bytes(img, fmt))
                except Exception:
                    failed_other.append(s_no)
            if on_progress:
                on_progress(i + 1, total)

    out_buf.seek(0)
    return out_buf.read(), failed_no_sno, failed_empty_vern, failed_other
