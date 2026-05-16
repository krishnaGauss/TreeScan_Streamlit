import csv
import io
import zipfile
from functools import lru_cache
from pathlib import Path

import openpyxl
import fitz
from PIL import Image, ImageDraw

from constants import (
    SERIF_FONT, DEVA_FONT, TABLE_FONT,
    HEADER_PATH, FOOTER_PATH, PAGE_W,
)
from drawing import (
    get_font, measure,
    wrap_text_mixed, draw_mixed_line, line_script_metrics,
)

BASE = Path(__file__).parent

# ── Palette ───────────────────────────────────────────────────────────────────
_FOREST_GREEN    = "#1F3D2E"         # heading text, values
_DARK_GREEN_RGB  = (31, 61, 46)      # icon circle border
_SAGE_RGB        = (182, 219, 123)   # icon circle fill (#b6db7b)
_SAGE_HEX        = "#6B8F70"         # hex equivalent of _SAGE_RGB
_ICON_FILL       = "#b6db7b"         # faint green fill for SVG icon paths
_LABEL_COLOR     = "#4E6E52"         # field label text
_VALUE_COLOR     = "#1A2E20"         # field value text
_DIV_COLOR       = (185, 212, 190)   # dividers
_RULE_COLOR      = (158, 193, 163)   # thin horizontal rules in heading

# ── Column field definitions: (csv_key, display_label, svg_file) ─────────────
#   key=None → combined Latitude / Longitude block
_COL1 = [
    ("S_No_",     "S. No.",          "leafy-green.svg"),
    ("Vernacular","Vernacular Name",  "search.svg"),
    ("Scientific","Scientific Name", "sprout.svg"),
]
_COL2 = [
    ("Family",    "Family",               "users.svg"),
    (None,        "Latitude / Longitude", "map-pin.svg"),
]
_COL3 = [
    ("Plant_Orig","Plant Origin",  "tree-pine.svg"),
    ("Remarks_",  "Remark",       "notebook-pen.svg"),
]
_COLUMNS = [_COL1, _COL2, _COL3]


# ── Asset loaders ─────────────────────────────────────────────────────────────
@lru_cache(maxsize=32)
def _load_svg_icon(svg_filename: str, size: int,
                   stroke_color: str = "#1F3D2E",
                   fill_color: str = "none") -> Image.Image:
    content = (BASE / svg_filename).read_text()
    content = content.replace('stroke="currentColor"', f'stroke="{stroke_color}"')
    content = content.replace('fill="none"', f'fill="{fill_color}"')
    doc  = fitz.open(stream=content.encode(), filetype="svg")
    page = doc[0]
    scale = size / 24.0
    pix  = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=True)
    img  = Image.frombytes("RGBA", [pix.width, pix.height], pix.samples)
    doc.close()
    return img


@lru_cache(maxsize=4)
def _load_header(width: int) -> Image.Image:
    img = Image.open(HEADER_PATH).convert("RGB")
    return img.resize((width, int(width * img.height / img.width)), Image.LANCZOS)


@lru_cache(maxsize=4)
def _load_footer(width: int) -> Image.Image:
    img = Image.open(FOOTER_PATH).convert("RGB")
    img = img.crop((0, 155, img.width, img.height))
    return img.resize((width, int(width * img.height / img.width)), Image.LANCZOS)


_IMAGE_EXTS = {".jpg", ".jpeg", ".png"}


def _img_to_pdf_bytes(img_bytes: bytes) -> bytes:
    img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="PDF", resolution=150)
    buf.seek(0)
    return buf.read()


def _pdf_to_images(pdf_bytes: bytes, width: int,
                   skip_footer_crop: bool = False) -> list[Image.Image]:
    doc    = fitz.open(stream=pdf_bytes, filetype="pdf")
    images = []
    for page in doc:
        scale = width / page.rect.width
        pix   = page.get_pixmap(matrix=fitz.Matrix(scale, scale), alpha=False)
        img   = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        if not skip_footer_crop:
            img = img.crop((0, 0, img.width, img.height - int(img.height * 0.07)))
        images.append(img)
    doc.close()
    return images


# ── Drawing primitives ────────────────────────────────────────────────────────
def _draw_outlined_circle_icon(canvas: Image.Image, cx: int, cy: int,
                               radius: int, icon: Image.Image) -> None:
    """Sage-green filled circle with dark-green border, white icon centred inside."""
    draw = ImageDraw.Draw(canvas)
    draw.ellipse(
        [cx - radius, cy - radius, cx + radius, cy + radius],
        fill=_SAGE_RGB, outline=_DARK_GREEN_RGB, width=2,
    )
    iw, ih = icon.size
    canvas.paste(icon, (cx - iw // 2, cy - ih // 2), mask=icon.split()[3])


def _spaced_text_width(draw: ImageDraw.ImageDraw, text: str, font,
                       spacing: int) -> int:
    total = sum(int(draw.textlength(ch, font=font)) for ch in text)
    total += spacing * max(0, len(text) - 1)
    return total


def _draw_spaced_text(draw: ImageDraw.ImageDraw, text: str, x: int, y: int,
                      font, fill: str, spacing: int, bold: bool = False) -> None:
    cx = x
    for ch in text:
        if bold:
            draw.text((cx + 1, y), ch, font=font, fill=fill)
        draw.text((cx, y), ch, font=font, fill=fill)
        cx += int(draw.textlength(ch, font=font)) + spacing


def _bold_text(draw: ImageDraw.ImageDraw, xy: tuple, text: str, font, fill: str) -> None:
    """Simulate bold by painting the glyph twice, shifted 1 px right."""
    x, y = xy
    draw.text((x + 1, y), text, font=font, fill=fill)
    draw.text((x,     y), text, font=font, fill=fill)


# ── Main section renderer ─────────────────────────────────────────────────────
def _render_info_section(row: dict, width: int) -> Image.Image:
    # Layout constants
    SEC_PAD_H    = 20    # left/right padding inside the section
    SEC_PAD_TOP  = 22    # top padding above heading
    SEC_PAD_BOT  = 5    # bottom padding below content
    HEADING_H    = 62    # increased to fit larger title
    HEAD_GAP     = 24    # gap between heading and column content
    COL_GAP      = 52    # horizontal gap between adjacent columns (divider centred here)
    ICON_DIAM    = 50    # diameter of each icon circle
    ICON_SZ      = 28    # SVG rasterise size (px)
    ICON_TXT_GAP = 16    # gap from circle right edge to text block
    LABEL_VAL_GAP= 6     # gap between label line and first value line
    ITEM_GAP     = 36    # uniform gap between every row in the col-1 grid
    H_DIV_INSET  = 12   # white-space inset on each side of the horizontal divider
    ORN_SZ       = 25    # ornament leaf icon size
    LETTER_SPC   = 4     # extra px between heading characters

    LABEL_SZ = 15
    VALUE_SZ = 19
    TITLE_SZ = 31    # +30% from 24

    label_font = get_font(LABEL_SZ, TABLE_FONT)
    value_font = get_font(VALUE_SZ, SERIF_FONT)
    deva_font  = get_font(VALUE_SZ, DEVA_FONT)
    title_font = get_font(TITLE_SZ, TABLE_FONT)

    lbl_asc, lbl_desc = label_font.getmetrics()
    val_asc, val_desc = value_font.getmetrics()
    lbl_h = lbl_asc + lbl_desc
    val_h = val_asc + val_desc

    # ── Column geometry ───────────────────────────────────────────────────────
    content_w = width - 2 * SEC_PAD_H
    col_w     = (content_w - 2 * COL_GAP) // 3
    txt_w     = max(60, col_w - ICON_DIAM - ICON_TXT_GAP)

    col_x = [
        SEC_PAD_H,
        SEC_PAD_H + col_w + COL_GAP,
        SEC_PAD_H + 2 * (col_w + COL_GAP),
    ]

    # ── Pre-compute item data for all columns ─────────────────────────────────
    probe = ImageDraw.Draw(Image.new("RGB", (max(txt_w, 1), 10)))

    col_items: list[list[tuple[str, str, list[str], int]]] = []
    for column in _COLUMNS:
        items = []
        for key, label, svg_file in column:
            if key is None:
                lat = row.get("Latitude",  "—")
                lon = row.get("Longitude", "—")
                val_lines = [lat, lon]
            else:
                val_lines = wrap_text_mixed(
                    probe, row.get(key, ""), deva_font, value_font, txt_w
                )
            n = len(val_lines)
            text_block_h = lbl_h + LABEL_VAL_GAP + n * val_h + max(0, n - 1) * 3
            item_h = max(ICON_DIAM, text_block_h)
            items.append((label, svg_file, val_lines, item_h))
        col_items.append(items)

    # ── Row-height model ──────────────────────────────────────────────────────
    # Col 1 drives row heights; cols 2/3 items anchor to col-1 icon centres.
    #   col-2/3 item-0 → col-1 row-0 centre (S.No level)
    #   col-2/3 item-1 → col-1 row-2 centre (Scientific Name level)
    # Horizontal divider sits at col-1 row-1 centre (Vernacular level).
    n_rows = max(len(col) for col in col_items)

    # Row heights driven solely by col 1 → uniform ITEM_GAP between col-1 icons
    row_heights = [col_items[0][r][3] if r < len(col_items[0]) else 0
                   for r in range(n_rows)]

    row_y = [0] * n_rows
    for r in range(1, n_rows):
        row_y[r] = row_y[r - 1] + row_heights[r - 1] + ITEM_GAP

    col1_h = sum(row_heights) + max(0, n_rows - 1) * ITEM_GAP

    # Canvas height must also fit cols-2/3 items which anchor to col-1 centres
    _rel_centers = [row_y[r] + row_heights[r] // 2 for r in range(n_rows)]
    max_col_h = col1_h
    for c in range(1, len(_COLUMNS)):
        for ri, (_, _, _, ih) in enumerate(col_items[c]):
            ref_r = min(ri * 2, n_rows - 1)
            bot = _rel_centers[ref_r] + ih // 2
            max_col_h = max(max_col_h, bot)

    # ── Canvas ────────────────────────────────────────────────────────────────
    section_h = SEC_PAD_TOP + HEADING_H + HEAD_GAP + max_col_h + SEC_PAD_BOT
    canvas = Image.new("RGB", (width, section_h), (255, 255, 255))
    draw   = ImageDraw.Draw(canvas)

    # ── Heading ───────────────────────────────────────────────────────────────
    TITLE = "BOTANICAL INFORMATION"
    probe2 = ImageDraw.Draw(Image.new("RGB", (width, 10)))
    title_w = _spaced_text_width(probe2, TITLE, title_font, LETTER_SPC)
    _, th, _, toy = measure(draw, TITLE, title_font)

    rule_y  = SEC_PAD_TOP + HEADING_H // 2
    title_x = (width - title_w) // 2
    title_y = SEC_PAD_TOP + (HEADING_H - th) // 2 - toy

    orn      = _load_svg_icon("leafy-green.svg", ORN_SZ, "#6B8F70")
    orn_flip = orn.transpose(Image.Transpose.FLIP_LEFT_RIGHT)
    ow, oh   = orn.size
    orn_gap  = 9

    left_orn_x  = title_x - orn_gap - ow
    right_orn_x = title_x + title_w + orn_gap
    orn_y       = rule_y - oh // 2

    draw.line([(SEC_PAD_H, rule_y), (left_orn_x - orn_gap, rule_y)],
              fill=_FOREST_GREEN, width=5)
    draw.line([(right_orn_x + ow + orn_gap, rule_y), (width - SEC_PAD_H, rule_y)],
              fill=_FOREST_GREEN, width=5)

    canvas.paste(orn,      (left_orn_x,  orn_y), mask=orn.split()[3])
    canvas.paste(orn_flip, (right_orn_x, orn_y), mask=orn_flip.split()[3])

    _draw_spaced_text(draw, TITLE, title_x, title_y,
                      title_font, _FOREST_GREEN, LETTER_SPC, bold=True)

    # ── Vertical column dividers (full content height) ────────────────────────
    content_top = SEC_PAD_TOP + HEADING_H + HEAD_GAP
    content_bot = content_top + max_col_h
    for i in range(2):
        div_x = col_x[i + 1] - COL_GAP // 2
        draw.line([(div_x, content_top), (div_x, content_bot)], fill=_DIV_COLOR, width=2)

    # ── Horizontal divider: cols 2+3, at the vertical centre of Vernacular row ──
    # Icon vertical centres for each row in col 1 (absolute Y)
    col1_centers = [content_top + row_y[r] + row_heights[r] // 2 for r in range(n_rows)]

    has_h_div = n_rows > 1 and any(len(col_items[c]) > 1 for c in range(1, len(_COLUMNS)))
    if has_h_div:
        hdiv_y = col1_centers[1]          # centre of the Vernacular (middle) row
        div1_x = col_x[1] - COL_GAP // 2  # vertical divider between col1 & col2
        div2_x = col_x[2] - COL_GAP // 2  # vertical divider between col2 & col3
        # Separate padded line for each right-hand column
        draw.line([(div1_x + H_DIV_INSET, hdiv_y), (div2_x - H_DIV_INSET, hdiv_y)],
                  fill=_DIV_COLOR, width=1)
        draw.line([(div2_x + H_DIV_INSET, hdiv_y), (width - SEC_PAD_H - H_DIV_INSET, hdiv_y)],
                  fill=_DIV_COLOR, width=1)

    # ── Column content ────────────────────────────────────────────────────────
    for col_idx, items in enumerate(col_items):
        cx = col_x[col_idx]
        for r, (label, svg_file, val_lines, item_h) in enumerate(items):
            n            = len(val_lines)
            text_block_h = lbl_h + LABEL_VAL_GAP + n * val_h + max(0, n - 1) * 3
            text_x       = cx + ICON_DIAM + ICON_TXT_GAP
            icon_cx      = cx + ICON_DIAM // 2

            if col_idx == 0:
                # Col 1: icon centred in its row zone, text centred in zone height
                icon_cy  = col1_centers[r]
                rh       = row_heights[r]
                text_top = content_top + row_y[r] + (rh - text_block_h) // 2
            else:
                # Cols 2/3: item-0 → row-0 centre (S.No level)
                #           item-1 → row-2 centre (Scientific level)
                ref_r    = r * 2
                icon_cy  = col1_centers[min(ref_r, n_rows - 1)]
                text_top = icon_cy - text_block_h // 2

            icon = _load_svg_icon(svg_file, ICON_SZ,
                                  stroke_color=_FOREST_GREEN, fill_color=_ICON_FILL)
            _draw_outlined_circle_icon(canvas, icon_cx, icon_cy, ICON_DIAM // 2, icon)

            _, _, lox, loy = measure(draw, label, label_font)
            _bold_text(draw, (text_x - lox, text_top - loy),
                       label, label_font, _LABEL_COLOR)

            val_top = text_top + lbl_h + LABEL_VAL_GAP
            for line in val_lines:
                line_asc, _ = line_script_metrics(line, deva_font, value_font)
                draw_mixed_line(draw, line, text_x + 1, val_top + line_asc,
                                deva_font, value_font, _VALUE_COLOR)
                draw_mixed_line(draw, line, text_x, val_top + line_asc,
                                deva_font, value_font, _VALUE_COLOR)
                val_top += val_h + 3

    return canvas


# ── PDF assembler ─────────────────────────────────────────────────────────────
def build_tree_pdf(pdf_bytes: bytes, row: dict,
                   skip_footer_crop: bool = False) -> bytes:
    width  = PAGE_W
    margin = (PAGE_W - 1160) // 2

    info_img    = _render_info_section(row, width - 2 * margin)
    info_canvas = Image.new("RGB", (width, info_img.height), "white")
    info_canvas.paste(info_img, (margin, 0))

    spacer = Image.new("RGB", (width, 12), "white")
    footer = _load_footer(width)
    parts  = ([_load_header(width), spacer, info_canvas]
              + _pdf_to_images(pdf_bytes, width, skip_footer_crop))

    body_h = sum(p.height for p in parts)
    canvas = Image.new("RGB", (width, body_h + footer.height - 5), "white")
    y = 0
    for p in parts:
        canvas.paste(p, (0, y))
        y += p.height
    canvas.paste(footer, (0, body_h - 5))

    buf = io.BytesIO()
    canvas.save(buf, format="PDF", resolution=150)
    buf.seek(0)
    return buf.read()


# ── Spreadsheet helpers ───────────────────────────────────────────────────────
def _read_rows(file_bytes: bytes, filename: str) -> list[dict]:
    if filename.lower().endswith(".xlsx"):
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
        ws      = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows    = [
            {headers[i]: (str(cell.value) if cell.value is not None else "")
             for i, cell in enumerate(row)}
            for row in ws.iter_rows(min_row=2)
        ]
        wb.close()
        return rows
    return list(csv.DictReader(io.StringIO(file_bytes.decode("utf-8-sig"))))


def process_pdf_zip(
    zip_bytes: bytes, csv_bytes: bytes, filename: str = "data.csv",
    on_progress=None, zip_type: str = "pdf",
) -> tuple[bytes, list[str], list[str]]:
    rows   = {r["S_No_"].strip(): r for r in _read_rows(csv_bytes, filename)}
    in_zip = zipfile.ZipFile(io.BytesIO(zip_bytes))

    is_images = zip_type == "images"
    all_names = [
        n for n in in_zip.namelist()
        if not n.endswith("/") and (
            Path(n).suffix.lower() in _IMAGE_EXTS if is_images else True
        )
    ]
    all_stems  = {Path(n).stem.strip() for n in all_names}
    xlsx_s_nos = set(rows.keys())

    _sort_key       = lambda x: (not x.isdigit(), int(x) if x.isdigit() else x)
    pdf_not_in_xlsx = sorted(all_stems - xlsx_s_nos, key=_sort_key)
    xlsx_not_in_pdf = sorted(xlsx_s_nos - all_stems, key=_sort_key)

    matched = [n for n in sorted(all_names) if Path(n).stem.strip() in rows]
    total   = len(matched)
    out_buf = io.BytesIO()

    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as out_zip:
        for i, name in enumerate(matched):
            s_no      = Path(name).stem.strip()
            raw_bytes = in_zip.read(name)
            pdf_bytes = _img_to_pdf_bytes(raw_bytes) if is_images else raw_bytes
            out_zip.writestr(
                f"{s_no}.pdf",
                build_tree_pdf(pdf_bytes, rows[s_no], skip_footer_crop=is_images),
            )
            if on_progress:
                on_progress(i + 1, total)

    out_buf.seek(0)
    return out_buf.read(), pdf_not_in_xlsx, xlsx_not_in_pdf
